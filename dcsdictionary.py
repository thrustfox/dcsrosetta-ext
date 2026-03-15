from luaparser import ast
from openpyxl import Workbook
from openpyxl import load_workbook

from dcsdeepl import DcsDeepLTranslator
import io
import re
import time

from dcsms import DcsMsTranslator
from lua_checker import *
from util import *




def tokenize_filter(filter1):
    # if filter1 is None or all white space
    if not filter1 or filter1.strip() == "":
        return []
    
    # tokenize and strip
    tokens = [token.strip() for token in filter1.split(',')]
    
    # remove empty string after strip
    return [token for token in tokens if token]


class DcsDictionary:
    translator = None
    replacer = None
    dump_lua = False
    dump_dict = False
    useTest = False
    useAggressive = True
    b_stop_process = False
    merge_to_org = False

    bilingualStartSym = '\n( '
    bilingualEndSym = ' )'
    startSym = '!@#$<'
    endSym = '!@#$>'
    header_rows=1

    @classmethod
    def set_translator(cls, translator):
        cls.translator = translator
    
    @classmethod
    def set_replacer(cls, replacer):
        cls.replacer = replacer
    
    @classmethod
    def from_file_path(cls, path, field_filter=None):
        dd = cls()
        if field_filter:
            dd.field_filter = field_filter
        with open(path, encoding='UTF-8') as f:
            lua_str = f.read()
        dd._load_dict(lua_str)
        return dd

    @classmethod
    def from_dict(cls, dict, field_filter=None):
        dd = cls()
        dd.dict = dict
        dd.lua_str = cls.to_lua(dict)
        if field_filter:
            dd.field_filter = field_filter
            print('_field_filter')
            #print(field_filter)
        return dd

    @classmethod
    def from_lua_str(cls, lua_str: str):
        dd = cls()
        dd._load_dict(lua_str)
        return dd

    def __init__(self):
        self.dict = None
        self.lua_str = None
        self.field_filter = lambda f: True
        self.desc_org = ''
        self.desc_en_org = ''

    def __eq__(self, other):
        return self.dict == other.dict

    def __ne__(self, other):
        return not self.__eq__(other)

    def _load_dict(self, lua_str):
        self.lua_str = lua_str
        lua = ast.parse(lua_str)
        fields = lua.body.body[0].values[0].fields
        if self.dump_lua:
            print('\n=== dump_lua ===')
            print(self.lua_str)
            print('\n')

        keys = [f.key.s for f in fields if self.field_filter(f)]
        values = [f.value.s for f in fields if self.field_filter(f)]
        raw_dict = dict(zip(keys, values))
        self.dict = {}
        for k, v in raw_dict.items():
            if v:
                if 'description' == k:
                    self.desc_org = v
                if 'description_EN' == k:
                    self.desc_en_org = v
                
                self.dict[k] = DcsDictionary.prefix_string(v)
            else:
                self.dict[k] = v
        if self.dump_dict:
            print('\n=== dump_dict before ===')
            print(self.dict)
            print('\n')

    def translate_item_by_item(self, from_lang, to_lang):
        """
        Translates a dictionary making one translation request per item.
        WARNING: Slow! Better use translate_whole. Kept in case the other method fails
        :return: a translated DcsDictionary
        """
        merged_dict = {}
        count = len(self.dict)
        i = 0
        for k, v in self.dict.items():
            i += 1
            if not i % 10:
                print(f'Translating {i} of {count}')
            if v:
                t = self.translator.translate(v, from_lang, to_lang)
                merged_dict[k] = t
            else:
                merged_dict[k] = v
        if self.dump_dict:
            print('\n== dump_dict after ==')
            print(merged_dict)
            print('\n')
        return self.from_dict(merged_dict)

    def dump_alist(self, alist):
        ndx = 0
        for item in alist:
            print('[%d] %s' % (ndx, item))
            ndx = ndx + 1

    @staticmethod
    def fix_odd_backslashes(text):
        pattern = r'\\+(?=[^\n"\'\\])'  # find sequential '\' not followed by newline, ", '
        
        def replacer(match):
            backslashes = match.group(0)
            if len(backslashes) % 2 == 1:  # add one if odd
                return backslashes + '\\'
            return backslashes
    
        return re.sub(pattern, replacer, text)
    
    @staticmethod
    def fix_even_backslashes(text):
        pattern = r'(\\*)(?=\n|"|\')'  # find '\' followed by newline, ", '
        
        def replacer(match):
            backslashes = match.group(0)

            if DcsDictionary.useAggressive == False:
                if len(backslashes) % 2 == 0:  # add one if even
                    return backslashes + '\\'
                return backslashes
            else:
                return '\\'  # leave just one
    
        return re.sub(pattern, replacer, text)
    
    @staticmethod
    def prefix_odd_backslashes(text):
        pattern = r'\\+(?=[^\n"\'\\])'  # find sequential '\' not followed by newline, ", '
                                        # make the length half
        def replacer(match):
            backslashes = match.group(0)
            return '\\' * int(len(backslashes) / 2)
    
        return re.sub(pattern, replacer, text)
    
    @staticmethod
    def prefix_even_backslashes(text):
        pattern = r'(\\*)(?=\n|"|\')'  # find '\' followed by newline, ", '
        return re.sub(pattern, '', text) # remove it
    
    @staticmethod
    def fix_eol_backslashes(text):
        return re.sub(r'\\+$', '', text) if not text.endswith('\n') else text  # remove sequential '\' at end of line

    @staticmethod
    def fix_string(str):
        # fix backslashes before write to lua
        str = DcsDictionary.fix_odd_backslashes(str)
        str = DcsDictionary.fix_even_backslashes(str)
        str = DcsDictionary.fix_eol_backslashes(str)
        return str

    @staticmethod
    def prefix_string(str):
        # remove escape char after read from lua
        str = DcsDictionary.prefix_odd_backslashes(str)
        str = DcsDictionary.prefix_even_backslashes(str)
        str = DcsDictionary.fix_eol_backslashes(str)
        return str

    def check_filter(self, k, filter_set):
        if k in filter_set:
            return False
        return True

    def make_filter_set(self, keys, filter1):
        filter_list = tokenize_filter(filter1)
        
        res = set()
        for k in keys:
            for f in filter_list:
                if f in k:
                    res.add(k)
        return res
    
    def translate_whole(self, from_lang, to_lang, useArray = True, split_size = 0, delay = 0, filter1 = '', bilingual = False, lua_detect=False):
        """
        Translates a dictionary making one translation function call
        WARNING: If it doesn't work use translate_item_by_item
        :return: a translated DcsDictionary
        """

        total_data_rows = 0
        processed_count = 0  # total processed rows (includes empty string)
        translated_count = 0  # actually translated rows
        skipped_count = 0  # empty rows
        filtered_count = 0  # filtered rows
        script_skipped_count = 0 # script detected rows
        script_skipped_rows = []
        
        filter_set = self.make_filter_set(self.dict.keys(), filter1)

        trans_values = None
        if useArray:
            #whole_texts = list(filter(lambda d: d, self.dict.values()))
            whole_texts = []
            whole_texts_key = []
            row_idx = self.header_rows
            total_data_rows = len(self.dict)
            
            for k, v in self.dict.items():
                row_idx = row_idx + 1
                processed_count = processed_count + 1
                if v is None or v.strip() == '':
                    skipped_count = skipped_count + 1
                    continue
                if not self.check_filter(k, filter_set):
                    filtered_count = filtered_count + 1
                    continue

                if lua_detect and is_lua_script(v):
                    script_skipped_count = script_skipped_count + 1
                    script_skipped_rows.append(row_idx)
                    continue
                translated_count = translated_count + 1
                whole_texts.append(v)
                whole_texts_key.append(k)
            if len(whole_texts) == 0:
                raise (Exception('no text to translate!'))
            print(f'Total {len(whole_texts)} items')

            if split_size == 0:
                start = time.time()
                trans_wholes = self.translator.translate(whole_texts, from_lang, to_lang)
                end = time.time()
                print(f'Total {len(whole_texts)} / {end - start:.5f} sec')
                trans_values = trans_wholes
            else:
                translated_chunks = []
                accu = 0
            
                for i in range(0, len(whole_texts), split_size):
                    chunk = whole_texts[i:i+split_size]
                    if not chunk:
                        continue
                    accu = accu + len(chunk)
                    if i > 0 and delay:
                        print(f'Sleeping {delay}s...')
                        time.sleep(delay)
                    if DcsDictionary.b_stop_process:
                        raise (Exception('Translation canceled!'))
                    start = time.time()
                    trans_chunk = self.translator.translate(chunk, from_lang, to_lang)
                    end = time.time()
                    print(f'Chunk {i // split_size + 1}: {accu} / {end - start:.5f} sec')
                    translated_chunks.extend(trans_chunk)
                
                trans_values = translated_chunks

        else:
            # Not used
            
            # Using less symbols seems to increase the probabilities of the translator being confused
            # Not useful. Kept as legacy method
            joinStr = '%s {} %s' % (self.startSym, self.endSym)
            whole_text = '. '.join(joinStr.format(v) for v in self.dict.values() if v)
            start = time.time()
            trans_whole = self.translator.translate(whole_text, from_lang, to_lang)
            end = time.time()
            print(f'Total {len(self.dict)} / {end - start:.5f} sec')
            trans_whole = trans_whole
            splitStr = '%s. %s' % (self.endSym, self.startSym)
            trans_values = [v.strip() for v in trans_whole.split(splitStr)]
            
            trans_values[0] = trans_values[0][(len(self.startSym) + 1):]
            trans_values[-1] = trans_values[-1][:-(len(self.endSym) + 1)]

        trans_values = self.replacer.replace_items(trans_values)
        trans_dict = dict(zip(whole_texts_key, trans_values))
        
        merged_dict = self.dict.copy()
        if bilingual:
            for k, v in trans_dict.items():
                if k in merged_dict:
                    merged_dict[k] = v + self.bilingualStartSym + merged_dict[k].rstrip() + self.bilingualEndSym
        else:
            merged_dict.update(trans_dict)
        
        #trans_values_iter = iter(trans_values)
        #for k, v in self.dict.items():
        #    if v:
        #        new_v = ''
        #        if self.check_filter(k, filter_set):
        #            next_v = next(trans_values_iter)
        #            if bilingual:
        #                new_v = next_v + self.bilingualStartSym + v.rstrip() + self.bilingualEndSym
        #            else:
        #                new_v = next_v
        #        else:
        #            new_v = v
        #        merged_dict[k] = new_v
        #    else:
        #        merged_dict[k] = ''
        if self.dump_dict:
            print('\n== dump_dict after ==')
            print(merged_dict)
            print('\n')
            
        print(f"\n=== Translation Complete ===")
        print(f"Total processed: {total_data_rows}")
        print(f"Translated: {translated_count}")
        print(f"Filter applied: {filtered_count}")
        print(f"Empty cells: {skipped_count}")
        if lua_detect == True:
            print(f"Scripts skipped: {script_skipped_count}")
            script_skipped_result = ", ".join(map(str, script_skipped_rows))
            print(f"Scripts skipped rows: {script_skipped_result}")
        print(f"")
            
        return self.from_dict(merged_dict)

    def save(self, dest_path, new_lua_str = None):
        with open(dest_path, 'w', encoding='UTF-8') as f:
            if new_lua_str == None:
                f.write(self.lua_str)
            else:
                f.write(new_lua_str)

    @classmethod
    def to_lua(cls, dict):
        strio = io.StringIO()
        strio.write('dictionary = \n')
        strio.write('{\n')
        for key, value in dict.items():
            value = DcsDictionary.fix_string(value.strip())
            #if value.endswith('\\'):  # This caused a huge amount of headaches!
            #    value = value[:-1]
            strio.write(f'    ["{key}"] = "{value}",\n')
        strio.write('} -- end of dictionary\n')
        strio.write('')
        return strio.getvalue()

    def save_to_xls(self, org_dict, filename):
        # create workbook
        wb = Workbook()
        ws = wb.active

        header = ['ID', 'Text', 'Translated Text']

        # append header
        ws.append(list(header))
        
        # append data
        for key, value in self.dict.items():
            if key in org_dict:
                if self.useTest:
                    ws.append([key, org_dict[key], '123' + value + '\n\n'])
                else:
                    ws.append([key, org_dict[key], value])
            else:
                ws.append([key, '', value])

        # auto-adjust columns width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)  # calculate max length
            adjusted_width = (max_length + 2) * 1.2  # append extra width
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 100)
        
        # save Excel
        wb.save(filename)

    def load_from_xls(self, org_dict, filename):
        # load from Excel
        wb = load_workbook(filename)
        ws = wb.active
        
        # create dictionary
        update_dict = {}
        
        # load data
        for row in ws.iter_rows(min_row=2, values_only=True):  # consider that first row is header
            key, _, value = row  # extract first col(key), third col(value)
            new_value = "" if value == None else value
            if key:
                if not isinstance(new_value, str):
                    new_value = str(new_value)
                update_dict[key] = new_value

        src_dict = None
        if self.merge_to_org:
            src_dict = org_dict
        else:
            src_dict = self.dict
        
        if check_keys_match(src_dict, update_dict) == False:
            print('\nWarning! Compatibility check failed. Mission generation is still available.')
            update_dict = merge_dicts(src_dict, update_dict)
        else:
            print('\nCompatibility check sucess!')
        
        if self.dump_dict:
            print('\n== dump_dict updated ==')
            print(update_dict)
            print('\n')
        return self.from_dict(update_dict)

    def get_replacer_data_text():
        if DcsDictionary.replacer != None:
            return DcsDictionary.replacer.get_data_text()
        return ''
    
